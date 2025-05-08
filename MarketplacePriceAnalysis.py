# Dependencies
import requests
import pandas as pd
import datetime
import openpyxl
import json

# Placeholder for API key (to be provided securely)
API_KEY = "YOUR_API_KEY"

# Get the current date and format it as a string
now = datetime.datetime.now()
timestamp = now.strftime("%Y-%m-%d")

# Placeholder list of product identifiers (e.g., UPCs or GTINs)
product_ids = ['SAMPLE_ID_1', 'SAMPLE_ID_2', 'SAMPLE_ID_3']  # Replace with actual IDs

# Fetch offers data using a web scraping API
offers = []
for product_id in product_ids:
    params = {
        'api_key': API_KEY,
        'type': 'offers',
        'domain': 'example.com',  # Generic domain
        'product_id': product_id,
        'max_page': 5
    }

    response = requests.get('https://api.example.com/request', params=params)  # Generic API endpoint
    data = response.json()
    offers.append(data)

# Save the offers list to a text file with a date stamp
file_name = f'{timestamp}_offers_list.txt'
with open(file_name, 'w') as file:
    json.dump(offers, file, indent=4)

# Initialize lists to store extracted values
product_ids_from_json = []
item_codes = []
prices = []
currencies = []
sellers = []
seller_ids = []
positions = []
pages = []
seller_links = []
condition_titles = []
condition_comments = []

# Iterate through the JSON data
for entry in offers:
    if "request_parameters" in entry:
        product_id = entry["request_parameters"]['product_id']
        for result in entry.get("offers", []):
            item_code = result.get("item_code")
            price = result.get("price", {}).get("value")
            currency = result.get("price", {}).get("currency")
            seller = result.get("seller", {}).get("name")
            seller_id = result.get("seller", {}).get("id")
            position = result.get("position")
            page = result.get("page")
            seller_link = result.get("seller", {}).get("link")
            condition_title = result.get("condition", {}).get("title")
            condition_comment = result.get("condition", {}).get("comments")

            product_ids_from_json.append(product_id)
            item_codes.append(item_code)
            prices.append(price)
            currencies.append(currency)
            sellers.append(seller)
            seller_ids.append(seller_id)
            positions.append(position)
            pages.append(page)
            seller_links.append(seller_link)
            condition_titles.append(condition_title)
            condition_comments.append(condition_comment)

# Create a DataFrame
data_df = pd.DataFrame({
    "ProductID": product_ids_from_json,
    "ItemCode": item_codes,
    "Price": prices,
    "Currency": currencies,
    "Seller": sellers,
    "SellerID": seller_ids,
    "Position": positions,
    "Page": pages,
    "SellerLink": seller_links,
    "ConditionTitle": condition_titles,
    "ConditionComment": condition_comments
})

# Convert ProductID to integer
data_df['ProductID'] = data_df["ProductID"].astype(int, errors='ignore')

# Load a price list CSV file (generic path)
price_list_df = pd.read_csv('resources/price_list.csv')  # Generic file path

# Clean and select relevant columns
price_list_df_cleaned = price_list_df.dropna(subset=['ProductID', 'MAPPrice'])
price_list_df_relevant = price_list_df_cleaned[['MaterialID', 'ProductID', 'Description', 'MAPPrice']]

# Merge the data and price list DataFrames
merged_df = pd.merge(data_df, price_list_df_relevant, on="ProductID", how="right")

# Compare prices
merged_df['PriceDifference'] = merged_df['Price'] - merged_df['MAPPrice']

# Sort by price difference descending
merged_df_sorted = merged_df.sort_values(by='PriceDifference', ascending=False)

# Reorder columns
merged_df_sorted = merged_df_sorted[['ProductID', 'ItemCode', 'Price', 'Currency', 'Seller', 'SellerID',
                                    'Position', 'Page', 'SellerLink', 'MaterialID', 'Description', 'MAPPrice',
                                    'PriceDifference', 'ConditionTitle', 'ConditionComment']]

# Export to Excel with a timestamp
file_name_2 = f'{timestamp}_marketplace_analysis.xlsx'
merged_df_sorted.to_excel(file_name_2, index=False)

# Load the Excel file and add hyperlinks to seller links
wb = openpyxl.load_workbook(file_name_2)
ws = wb.active
seller_link_index = merged_df_sorted.columns.get_loc('SellerLink') + 1

# Convert SellerLink column to hyperlinks
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=seller_link_index, max_col=seller_link_index):
    for cell in row:
        if cell.value:
            cell.hyperlink = cell.value
            cell.style = 'Hyperlink'

# Save the updated workbook
wb.save(file_name_2)